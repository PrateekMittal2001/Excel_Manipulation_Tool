import openpyxl as op
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import get_column_letter
import pandas as pd
import numpy as np


class ExcelManipulations:

    @staticmethod
    def copy():
        wb = Workbook()

        # Output File
        sheet1 = wb.active
        sheet1.title = 'Student List'
        sheet1.cell(column=1, row=1).value = 'Student List'
        studentlist = [('RollNo', 'Name', 'age', 'marks'), (1, 'Rahu', 30, 50),
                       (2, 'sharma', 40, 60), (3, 'Ketu', 44, 75)]
        for col in range(1, 5):
            for row in range(1, 5):
                sheet1.cell(column=col, row=1 + row).value = studentlist[row - 1][col - 1]
        wb.save('C:\\Users\\HP\\Excel_Exp\\CopyOutput.xlsx')

        # Input Sheet
        sheet1 = wb.active
        sheet1.title = 'Student List'
        sheet1.cell(column=1, row=1).value = 'Student List'
        studentlist = [('RollNo', 'Name', 'age', 'marks'), (1, 'Juhi', 20, 100),
                       (2, 'dilip', 20, 110), (3, 'jeevan', 24, 145)]
        for col in range(1, 5):
            for row in range(1, 5):
                sheet1.cell(column=col, row=1 + row).value = studentlist[row - 1][col - 1]
        wb.save('C:\\Users\\HP\\Excel_Exp\\students.xlsx')
        sheet1['B6'] = 'Average'
        sheet1['C6'] = '=AVERAGE(C3:C5)'
        sheet1['D6'] = Translator('=AVERAGE(C3:C5)', origin="C6").translate_formula("D6")
        wb.save('C:\\Users\\HP\\Excel_Exp\\students.xlsx')

        print("You want to copy 1)complete column 2)single element")
        chh = int(input())

        if chh == 1:
            wb3 = Workbook()
            sheet3 = wb3.active
            print("How many columns you want to copy: ")
            number = int(input())
            j = 1
            for q in range(number):
                print("Enter the column number you want to copy: ")
                colls = int(input())
                for i in range(1, sheet1.max_row + 1):
                    c = sheet1.cell(row=i, column=colls)
                    sheet3.cell(row=i, column=j).value = c.value
                j = j + 1
            wb3.save('C:\\Users\\HP\\Excel_Exp\\CopyColumn.xlsx')

        elif chh == 2:
            # DATA COPY
            print("Enter the element's row and column you want to copy")
            rowi = int(input())
            columni = int(input())
            ele = sheet1.cell(rowi, columni).value
            # print(ele)

            wb1 = op.load_workbook("C:\\Users\\HP\\Excel_Exp\\CopyOutput.xlsx")
            sheet2 = wb1.active
            sheet2.title = "COPY"
            sheet2.cell(row=rowi, column=columni).value = ele

            wb1.save("C:\\Users\\HP\\Excel_Exp\\CopyOutput.xlsx")

            # FORMULA COPY
            print("Enter the Formula's row and column you want to copy")
            rowi = int(input())
            columni = int(input())
            ele = sheet1.cell(rowi, columni).value
            # print(ele)

            wb1 = op.load_workbook("C:\\Users\\HP\\Excel_Exp\\CopyOutput.xlsx")
            sheet2 = wb1.active
            sheet2.title = "COPY"
            sheet2.cell(row=rowi, column=columni).value = ele

            wb1.save("C:\\Users\\HP\\Excel_Exp\\CopyOutput.xlsx")

        else:
            print("Please enter a valid choice.")

    @staticmethod
    def highlight():

        wb = op.load_workbook("C:\\Users\\HP\\Excel_Exp\\Highlight.xlsx")
        sheet = wb.active

        for i in range(3, 6):
            ele = sheet.cell(row=i, column=4).value
            if ele < 40:
                sheet['D3'].fill = PatternFill("solid", fgColor="F50707")
            elif ele < 75:
                sheet['D4'].fill = PatternFill("solid", fgColor="FFBF00")
            else:
                sheet['D5'].fill = PatternFill("solid", fgColor="71FF33")
        wb.save("C:\\Users\\HP\\Excel_Exp\\HighlightOutput.xlsx")

    @staticmethod
    def filter():

        wb = Workbook()
        sheet = wb.active

        data = [
            ["Fruits", "Quantity"],
            ["Kiwi", 3],
            ["Grape", 15],
            ["Apple", 5],
            ["Peach", 6],
            ["Pomegranate", 7],
            ["Pear", 9],
            ["Tangerine", 3],
            ["Blueberry", 3],
            ["Mango", 3],
            ["Watermelon", 3],
            ["Blackberry", 3],
            ["Orange", 3],
            ["Raspberry", 3],
            ["Banana", 3]
        ]
        for i in data:
            sheet.append(i)

        sheet.auto_filter.ref = "A1:B15"
        sheet.auto_filter.add_filter_column(0, ["Kiwi", "Apple", "Mango"])
        sheet.auto_filter.add_sort_condition("B2:B15")

        wb.save("C:\\Users\\HP\\Excel_Exp\\filtered.xlsx")

    @staticmethod
    def insert():

        print("Enter where would you like to insert the text (1.Specific Cell 2.Whole Row 3.Whole Column)")
        choice = int(input())

        wb = op.load_workbook("C:\\Users\\HP\\Excel_Exp\\InsertText.xlsx")
        wb1 = op.load_workbook("C:\\Users\\HP\\Excel_Exp\\InsertText.xlsx")
        wb2 = op.load_workbook("C:\\Users\\HP\\Excel_Exp\\InsertText.xlsx")
        sheet = wb.active
        sheet2 = wb1.active
        sheet3 = wb2.active

        if choice == 1:

            c1 = sheet.cell(row=7, column=1)
            c1.value = "4"

            c2 = sheet.cell(row=7, column=2)
            c2.value = "Rahul"

            print("Enter the Age: ")
            age = int(input())
            print("Enter the Marks:")
            marks = int(input())

            c3 = sheet.cell(row=7, column=3)
            c3.value = int(age)
            c3 = sheet.cell(row=7, column=4)
            c3.value = marks

            wb.save("C:\\Users\\HP\\Excel_Exp\\InsertTextCell.xlsx")

        elif choice == 2:

            print("Enter the row: ")
            rows = int(input())
            cols = sheet2.max_column
            print("Enter the text: ")
            text = input()
            for i in range(cols):
                sheet2.cell(row=rows, column=i + 1).value = text

            wb1.save("C:\\Users\\HP\\Excel_Exp\\InsertTextRow.xlsx")

        elif choice == 3:

            print("Enter the column: ")
            cols = int(input())
            rows = sheet3.max_row
            print("Enter the text: ")
            text = input()
            for i in range(rows):
                sheet3.cell(row=i + 1, column=cols).value = text

            wb2.save("C:\\Users\\HP\\Excel_Exp\\InsertTextColumn.xlsx")

        else:
            print("Please enter a valid choice.")

        # sheet.insert_rows(7)
        # sheet.insert_cols(7)

        wb.save("C:\\Users\\HP\\Excel_Exp\\InsertText.xlsx")

    @staticmethod
    def replace():

        wb = op.load_workbook("C:\\Users\\HP\\Excel_Exp\\Replace.xlsx")
        sheet = wb.active

        replacement = {'pass': 'yes', 'fail': 'noo'}

        print("Enter replacement you want(1.Whole Sheet 2.Specific Row 3.Specific Column)")
        choice = int(input())
        if choice == 1:
            # For Whole Sheet

            number_rows = sheet.max_row
            number_columns = sheet.max_column

            for i in range(number_columns):
                for k in range(number_rows):
                    cell = str(sheet[get_column_letter(i + 1) + str(k + 1)].value)
                    for key in replacement.keys():
                        if str(cell) == key:
                            newcell = replacement.get(key)
                            sheet[get_column_letter(i + 1) + str(k + 1)] = str(newcell)
            wb.save('C:\\Users\\HP\\Excel_Exp\\ReplaceWholeSheet.xlsx')

        elif choice == 2:
            # For Specific Row

            print("Please enter the row: ")
            number_rows = int(input())
            number_columns = sheet.max_column

            for i in range(number_columns):
                k = number_rows - 1
                cell = str(sheet[get_column_letter(i + 1) + str(k + 1)].value)
                for key in replacement.keys():
                    if str(cell) == key:
                        newcell = replacement.get(key)
                        sheet[get_column_letter(i + 1) + str(k + 1)] = str(newcell)
            wb.save('C:\\Users\\HP\\Excel_Exp\\ReplaceRow.xlsx')

        elif choice == 3:
            # For Specific Column

            print("Please enter the column: ")
            number_rows = sheet.max_row
            number_columns = int(input())

            i = number_columns - 1
            for k in range(number_rows):
                cell = str(sheet[get_column_letter(i + 1) + str(k + 1)].value)
                for key in replacement.keys():
                    if str(cell) == key:
                        newcell = replacement.get(key)
                        sheet[get_column_letter(i + 1) + str(k + 1)] = str(newcell)
            wb.save('C:\\Users\\HP\\Excel_Exp\\ReplaceColumn.xlsx')

        else:
            print("Please Enter a valid choice")

    @staticmethod
    def pivot():

        """df = pd.read_excel("C:\\Users\\HP\\Excel_Exp\\Weather.xlsx")
        print(df)

        print(df.pivot(index='city', columns='date'))
        print(df.pivot(index='city', columns='date', values="humidity"))
        print(df.pivot(index='humidity', columns='city'))
        print(df.pivot(index='date', columns='city'))

        writer = pd.ExcelWriter('C:\\Users\\HP\\Excel_Exp\\PivotOutput.xlsx')

        table = pd.pivot_table(df, index='humidity', columns='city')
        for i in table.index.get_level_values(0).unique():
            temp_df = table.xs(i, level=0)
            temp_df.to_excel(writer, i)
        writer.save()
        """
        df = pd.read_excel("C:\\Users\\HP\\Excel_Exp\\pivotsales.xlsx")
        table = pd.pivot_table(df, index=["Manager", "Rep", "Product"], values=["Price", "Quantity"],aggfunc=[np.sum, np.mean], fill_value=0)
        # for manager in table.index.get_level_values(0).unique():
        #     print(table.xs(manager, level=0))

        writer = pd.ExcelWriter('pivotOutput.xlsx')
        for manager in table.index.get_level_values(0).unique():
            temp_df = table.xs(manager, level=0)
            temp_df.to_excel(writer, manager)
        writer.save()

    @staticmethod
    def formula():

        wb = op.load_workbook("C:\\Users\\HP\\Excel_Exp\\Formula.xlsx")
        sheet = wb.active

        sheet['C6'] = "=AVERAGE(C3:C5)"
        sheet['D6'] = "=AVERAGE(D3:D5)"
        sheet['C7'] = '=SUM(C3:C5)'
        sheet['D7'] = '=SUM(D3:D5)'

        # Column's Formula
        print("Enter the Cell: ")
        cell = input()
        print("Enter the Formula: ")
        formulae = input()
        sheet[cell] = formulae

        wb.save("C:\\Users\\HP\\Excel_Exp\\formulaOutput.xlsx")

    @staticmethod
    def find():

        wb = op.load_workbook("C:\\Users\\HP\\Excel_Exp\\Find.xlsx")
        sheet = wb.active

        print("Enter the string to be found: ")
        inputstring = input()

        rows = sheet.max_row
        cols = sheet.max_column

        fr = []
        fc = []
        k = 0

        for i in range(0, 100):
            fc.append(0)
            fr.append(0)

        for i in range(0, rows):
            for j in range(0, cols):
                if sheet.cell(row=i + 1, column=j + 1).value == inputstring:
                    fr[k] = i + 1
                    fc[k] = j + 1
                    k += 1

        if fc[0] == 0 and fr[0] == 0:
            print("String not found")
        else:
            print("String Found")

            print("Do you want to 1.Replace Text 2.Highlight 3.Insert row")
            inputchoice = int(input())

            if inputchoice == 1:
                print("Enter the string you want to replace with")
                stringchange = input()
                for i in range(0, k):
                    sheet.cell(row=fr[i], column=fc[i]).value = stringchange

            elif inputchoice == 2:
                print("Enter the color: 1.Red 2.yellow 3.Green")
                color = int(input())
                for i in range(0, k):
                    if color == 1:
                        sheet.cell(row=fr[i], column=fc[i]).fill = PatternFill("solid", fgColor="F50707")
                    elif color == 2:
                        sheet.cell(row=fr[i], column=fc[i]).fill = PatternFill("solid", fgColor="FFBF00")
                    elif color == 3:
                        sheet.cell(row=fr[i], column=fc[i]).fill = PatternFill("solid", fgColor="71FF33")
                    else:
                        print("Please enter a valid choice.")
            elif inputchoice == 3:
                sheet.insert_rows(rows)

        wb.save("C:\\Users\\HP\\Excel_Exp\\FindOutput.xlsx")


x1 = ExcelManipulations()
x1.formula()
